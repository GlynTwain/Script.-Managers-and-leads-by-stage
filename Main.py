import datetime
import os
import string
import openpyxl
from openpyxl.styles import Alignment
import win32com.client as win32

"""pyinstaller --onefile main.py """

name_colum_one = 0
chop = 0
lids_score = 1
bib = 1
sheet = 0
workbook = 0
name_new_file = "report.xlsx"

number_of_sources = {}

dis_stades = {
    "Всего лидов": 0,
    "Не обработан": 0,
    "Связь не установлена": 0,
    "Принят в работу": 0,
    "Сбор документов": 0,
    "Конвертация в сделку": 0,
    "Отказ клиентом": 0,
    "Некачественный лид": 0,
    "Дубль": 0,
    "Неактуально": 0,
    "Отказ клиенту": 0,
    "Сотрудничество": 0,
    "Соискатель": 0,
    "Клиент (уточнение информации)": 0,
    "Отказ ЛПР": 0

}


def File_convert():
    global sheet
    global workbook
    global name_new_file
    file = str(os.path.abspath(os.curdir)) + "\\report.xlsx"
    file_old = str(os.path.abspath(os.curdir)) + "\\report.xls"

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wbo = excel.Workbooks.Open(file_old)

    wbo.SaveAs(file, FileFormat=51)
    # FileFormat = 51 is for .xlsx extension
    # FileFormat = 56 is for .xls extension
    wbo.Close()
    del wbo
    excel.Application.Quit()

    workbook = openpyxl.load_workbook(filename=name_new_file)
    sheet = workbook.active

    Start()


def Appraiser():
    nuuumer = 0

    while True:
        nuuumer = nuuumer + 1
        this_sources = sheet.cell(row=nuuumer, column=1).value
        global lids_score
        if this_sources != sheet.cell(row=999, column=1).value:
            lids_score = lids_score + 1

        if this_sources == sheet.cell(row=999, column=1).value:
            break


def Sources_and_Sum():
    global name_colum_one
    name_colum_one = sheet.cell(row=1, column=1).value

    for i in range(2, lids_score):

        this_cell = sheet.cell(row=i, column=1).value

        if not this_cell in number_of_sources:
            number_of_sources[this_cell] = 0

        if this_cell in number_of_sources:
            number_of_sources[this_cell] += 1


def Convertering():
    for key in number_of_sources:
        buferr = number_of_sources[key]
        number_of_sources[key] = dict(dis_stades)
        stringls = key
        number_of_sources[stringls]["Всего лидов"] = buferr


def Stages():
    for num in range(2, lids_score):
        key_sources = sheet.cell(row=num, column=1).value
        key_stages = sheet.cell(row=num, column=2).value
        number_of_sources[key_sources][key_stages] += 1


def List_Creator():
    global sheet
    global workbook
    workbook.create_sheet(index=1, title="Отчёт")
    sheet = workbook["Отчёт"]


def CompletionFirst():
    pip = 1
    sheet.cell(row=1, column=pip).value = name_colum_one

    for _stages in dis_stades:
        pip = pip + 1
        sheet.cell(row=1, column=pip).value = _stages


def CompletionTwo():
    global bib
    for key in number_of_sources:
        bib = bib + 1
        num = 1
        sheet.cell(row=bib, column=1).value = key
        for stades in number_of_sources[key]:
            num = num + 1
            sheet.cell(row=bib, column=num).value = number_of_sources[key][stades]


def Format():
    for i in range(1, 31):
        sheet.row_dimensions[i].height = 20

    for c in string.ascii_letters:
        sheet.column_dimensions[c].width = 20
    sheet.column_dimensions['A'].width = 46

    for r in range(1, 50):

        for w in range(1, 50):
            vino = sheet.cell(row=r, column=w)
            vino.alignment = Alignment(horizontal='center')


def Saved():
    global workbook
    direction = 0
    if name_colum_one == "Ответственный":
        direction = "Менеджерам и лидам"
    if name_colum_one == "Источник лида":
        direction = "Источникам и лидам"

    workbook.save("Отчёт по " + direction + " (" + modification_date(name_new_file) + ").xlsx")
    os.remove(name_new_file)


def modification_date(filename):
    """ Записывает дату после создания файла по его свойствам, по свежему так сказать"""
    t = os.path.getmtime(filename)
    return datetime.datetime.fromtimestamp(t).strftime("%d.%m")


def Sumer():
    for sources in number_of_sources:
        for stadiss in number_of_sources[sources]:
            dis_stades[stadiss] = dis_stades[stadiss] + number_of_sources[sources][stadiss]
    global bib
    bib = bib + 1
    i = 1
    sheet.cell(row=bib, column=1).value = "ИТОГО: "
    for key in dis_stades:
        i = i + 1
        sheet.cell(row=bib, column=i).value = dis_stades[key]


def Start():
    Appraiser()
    Sources_and_Sum()
    Convertering()
    Stages()
    List_Creator()
    CompletionFirst()
    CompletionTwo()
    Sumer()
    Format()
    print(number_of_sources)
    Saved()


for i in range(1, 2):
    File_convert()
